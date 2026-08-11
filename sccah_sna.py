#!/usr/bin/env python3
"""
SCCAH — Knowledge Graph + Social Network Analysis
==================================================
Companion analysis script for the coding workbook 'SCCAH_Workbook.xlsx'
(Knowledge Graph and Social Network Analysis of Surveillance-Capitalist
Capabilities and Autonomy Harms).

WHAT THIS SCRIPT DOES:

  1. LOAD      Place 'SCCAH_Workbook*.xlsx' in the SAME FOLDER as this
               script (so you can keep updating the workbook and simply
               re-run). Reads three sheets:
                 - Evidence_Items          (one row per coded claim)
                 - Sources_Log             (joined on Source_ID for the
                                            publication Year + citation)
                 - Controlled_Vocabulary   (node names / definitions)

  2. FILTER    Keeps only rows where Include == 'include' AND all four
               harm criteria are 'Y' (the binary inclusion gate, s3.4 of
               the Methods / s4.1 of the Coding Protocol). No date
               filtering is applied — studies of ALL ages are counted.

  3. GRAPH     Builds the weighted bipartite network: capability nodes x
               autonomy-dimension nodes. Edge weight = SUM of evidential
               status scores (1/2/3) over the independent evidence items
               supporting that edge (w = Σ e_i; s3.8). The independent-
               claims rule is enforced upstream in the workbook itself
               (corroborating sources share one row), so each row = one
               independent item. If any edge accumulates > 10 items the
               script flags it and additionally reports log-scaled
               weights w = log(1 + Σ e_i), per the protocol (s2.3).

  4. SNA       Runs the five sensitivity passes of s3.9:
                 Pass A  full weighted graph (baseline)
                 Pass B  empirical-only (evidential status 1 excluded)
                 Pass C  binary graph (all edge weights = 1)
                 Pass D  source-type-filtered graphs (one per type)
                 Pass E  Opsahl generalised degree, alpha swept over
                         {0, 0.25, 0.5, 0.75, 1}
               The primary measure is WEIGHTED DEGREE centrality
               (stable under weight perturbation — Borgatti et al. 2006;
               Segarra & Ribeiro 2015). Betweenness is reported as
               secondary description only, with that caveat. Spearman
               rank correlations between capability rankings quantify
               robustness (rho > 0.85 treated as robust). Bipartite
               community detection (greedy modularity maximisation on
               the weighted two-mode graph, approximating Barber 2007)
               is ran - MUST have networkx is installed.

  5. OUTPUT    - Console summary (rankings + Spearman matrix)
               - CSV files in ./sna_outputs/ (evidence table, edge list,
                 node metrics per pass, Spearman matrix, communities)
               - ./sna_outputs/SCCAH_network.html — a fully interactive,
                 self-contained visualisation (open in any browser):
                   * every evidence item is ONE line (scroll to zoom /
                     drag to pan: zoomed in, overlapping lines separate
                     and each is individually clickable)
                   * line COLOUR   = Eedge type
                   * line THICKNESS = evidential status (1/2/3)
                   * line BRIGHTNESS = publication year (brighter = newer)
                   * three views: the bipartite network, and two one-mode
                     projections — CAPABILITY (capabilities linked when
                     they bear on the same autonomy dimension) and
                     AUTONOMY (dimensions linked when affected by the
                     same capability); link weight = sum over shared
                     partners of the smaller of the two nodes' evidence
                     weights on that partner
                   * clean labels by default; a 'detailed mode' toggle
                     adds IDs (CAP-003, AUT-OPP...) and reveals the Pass A
                     ranking, Spearman sensitivity matrix and communities
                   * side-panel toggles: edge group, evidential
                     status, year slider; capability / dimension /
                     source-type filters live under 'Advanced filters'
                   * click a node  -> focus + details (definition,
                     metrics, connected edges)
                   * click a line  -> full evidence-item provenance
                     (citation, year, quote, mechanism, locator...)

DEPENDENCIES:  pandas, openpyxl   (pip install pandas openpyxl)
   optional:   networkx           (community detection + betweenness;
                                   everything else runs without it)

USAGE:         python sccah_sna.py
               (from anywhere — the workbook is located relative to the
               script file itself, not the working directory)
"""

from __future__ import annotations

import glob
import html
import json
import math
import os
import re
import sys
from collections import defaultdict
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

# ----------------------------------------------------------------------------
# 0. Configuration
# ----------------------------------------------------------------------------

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
WORKBOOK_PATTERN = "SCCAH_Workbook*.xlsx"   # workbook lives beside this script
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "sna_outputs")

ROBUSTNESS_RHO = 0.85          # Spearman threshold treated as "robust" (s3.9)
LOG_SCALE_TRIGGER = 10         # >10 items on one edge -> also report log scale
OPSAHL_ALPHAS = [0.0, 0.25, 0.5, 0.75, 1.0]

# Fixed hues (HSL) per superordinate edge-type group; brightness is then
# modulated by publication year in the visualisation.
GROUP_HUES = {
    "Facilitation":      145,   # green
    "Intensification":    28,   # amber/orange
    "Concealment/Power": 268,   # purple
}


# ----------------------------------------------------------------------------
# 1. Workbook loading & parsing
# ----------------------------------------------------------------------------


# Locating Workbook 
def find_workbook() -> str:
    candidates = [
        p for p in glob.glob(os.path.join(SCRIPT_DIR, WORKBOOK_PATTERN))
        if not os.path.basename(p).startswith("~$")        # skip Excel locks
    ]
    if not candidates:
        sys.exit(
            f"ERROR: no file matching '{WORKBOOK_PATTERN}' found in "
            f"{SCRIPT_DIR}\nPlace the coding workbook in the same folder "
            f"as this script."
        )
    candidates.sort(key=os.path.getmtime, reverse=True)
    if len(candidates) > 1:
        print(f"NOTE: {len(candidates)} matching workbooks found; using the "
              f"most recently modified: {os.path.basename(candidates[0])}")
    return candidates[0]


def sheet_rows(wb, name: str) -> list[tuple]:
    if name not in wb.sheetnames:
        sys.exit(f"ERROR: expected sheet '{name}' not found in workbook "
                 f"(sheets present: {wb.sheetnames})")
    return list(wb[name].iter_rows(values_only=True))


# Turning raw sheet rows into a DataFrame
def rows_to_df(rows: list[tuple], header_key: str) -> pd.DataFrame:
    header_idx = next(
        (i for i, r in enumerate(rows)
         if r and str(r[0]).strip() == header_key),
        None,
    )
    if header_idx is None:
        sys.exit(f"ERROR: could not find header row starting with ")
    header = [str(c).strip() if c is not None else f"col{j}"
              for j, c in enumerate(rows[header_idx])]
    body = [r for r in rows[header_idx + 1:] if r and r[0] not in (None, "")]
    df = pd.DataFrame(body, columns=header[: len(rows[header_idx])])
    # placeholder/template rows like "[ADD]" or "EV-XXXX" are dropped
    df = df[~df.iloc[:, 0].astype(str).str.startswith("[")]
    return df


# Pulling node names and definitions for capabilities and autonomy dimensions out of vocabulary sheet.
def parse_vocabulary(rows: list[tuple]) -> dict:
    vocab = {"cap": {}, "dim": {}, "edge": {}, "srctype": []}
    for r in rows:
        c0 = str(r[0]).strip() if r and r[0] is not None else ""
        if re.fullmatch(r"CAP-\d+", c0):
            vocab["cap"][c0] = {
                "name": str(r[1] or c0).strip(),
                "definition": str(r[2] or "").strip(),
            }
        elif re.fullmatch(r"AUT-[A-Z]+", c0):
            vocab["dim"][c0] = {
                "name": str(r[1] or c0).strip(),
                "definition": str(r[2] or "").strip(),
            }
        elif re.fullmatch(r"SRC-[A-Z]+", c0):
            vocab["srctype"].append(str(r[1] or "").strip())
        elif re.fullmatch(r"E\d{2}", c0):
            vocab["edge"][c0] = {
                "name": str(r[1] or c0).strip(),
                "group": str(r[2] or "").strip(),
                "rule": str(r[3] or "").strip(),
            }
    return vocab


# Evidence rows and Source log extracted and attached to publication year and full citation
def load_data(path: str):
    wb = load_workbook(path, read_only=True, data_only=True)

    vocab = parse_vocabulary(sheet_rows(wb, "Controlled_Vocabulary"))
    src = rows_to_df(sheet_rows(wb, "Sources_Log"), "Source_ID")
    ev = rows_to_df(sheet_rows(wb, "Evidence_Items"), "Evidence_ID")

    # Normalising the columns relied on 
    def col(df, *names):
        for n in names:
            for c in df.columns:
                if str(c).lower().startswith(n.lower()):
                    return c
        sys.exit(f"ERROR: none of the columns {names} found "
                 f"(available: {list(df.columns)})")

    year_col = col(src, "Year")
    cite_col = col(src, "Full citation")
    stype_src_col = col(src, "Source type")
    src_small = src[[src.columns[0], cite_col, year_col, stype_src_col]].copy()
    src_small.columns = ["Source_ID", "Citation", "Year", "SourceTypeLog"]
    src_small["Source_ID"] = src_small["Source_ID"].astype(str).str.strip()
    src_small["Year"] = pd.to_numeric(src_small["Year"], errors="coerce")

    keep = {
        "Evidence_ID": col(ev, "Evidence_ID"),
        "Source_ID": col(ev, "Source_ID"),
        "Capability": col(ev, "Capability_ID"),
        "Dimension": col(ev, "Autonomy_Dim_ID"),
        "EdgeFine": col(ev, "Edge type (fine"),
        "EdgeGroup": col(ev, "Edge type (super"),
        "Crit1": col(ev, "Capacity impact"),
        "Crit2": col(ev, "Social mediation"),
        "Crit3": col(ev, "Preventability"),
        "Crit4": col(ev, "Mechanism present"),
        "Include": col(ev, "Include"),
        "Status": col(ev, "Evidential status"),
        "SourceType": col(ev, "Source type"),
        "UK": col(ev, "UK relevance"),
        "Confidence": col(ev, "Coder confidence"),
        "Locator": col(ev, "Page / locator", "Page"),
        "Quote": col(ev, "Quote or paraphrase", "Quote"),
        "Corrob": col(ev, "Corroborating"),
        "Mechanism": col(ev, "Harm mechanism"),
        "Target": col(ev, "Target / affected party (who)", "Target"),
    }
    e = ev[[c for c in keep.values()]].copy()
    e.columns = list(keep.keys())
    for c in ("Evidence_ID", "Source_ID", "Capability", "Dimension",
              "EdgeFine", "EdgeGroup", "Include"):
        e[c] = e[c].astype(str).str.strip()

    n_rows = len(e)

    # Confirming harm criteria for each evidence item
    crit_ok = (
        e[["Crit1", "Crit2", "Crit3", "Crit4"]]
        .apply(lambda s: s.astype(str).str.strip().str.upper() == "Y")
        .all(axis=1)
    )
    included = e[(e["Include"].str.lower() == "include") & crit_ok].copy()
    n_excluded = n_rows - len(included)

    # attaching year + citation from Sources_Log
    included = included.merge(src_small, on="Source_ID", how="left")
    missing_year = included["Year"].isna().sum()
    if missing_year:
        print(f"WARNING: {missing_year} evidence item(s) have no Year in "
              f"Sources_Log; they will show as 'year unknown' and are "
              f"treated as the OLDEST year for the brightness scale.")

    included["Status"] = pd.to_numeric(included["Status"], errors="coerce")
    bad_status = included["Status"].isna() | ~included["Status"].isin([1, 2, 3])
    if bad_status.any():
        print(f"WARNING: {bad_status.sum()} item(s) with evidential status "
              f"outside {{1,2,3}} — dropped: "
              f"{included.loc[bad_status, 'Evidence_ID'].tolist()}")
        included = included[~bad_status]
    included["Status"] = included["Status"].astype(int)

    def canon_stype(v):
        v = str(v or "").strip()
        if not v or v.lower() == "nan":
            return ""
        for c in vocab["srctype"]:
            cl, vl = c.lower(), v.lower()
            if cl == vl or cl.startswith(vl) or vl.startswith(cl):
                return c
        return v
    fixed = included["SourceType"].map(canon_stype)
    changed = (fixed != included["SourceType"].astype(str).str.strip()).sum()
    if changed:
        print(f"NOTE: normalised {changed} source-type label(s) against the "
              f"Controlled_Vocabulary (e.g. 'Policy / NGO' -> "
              f"'Policy / NGO report').")
    included["SourceType"] = fixed

    print(f"Loaded {n_rows} evidence rows -> {len(included)} pass the "
          f"four-criterion inclusion gate ({n_excluded} excluded).")
    return included.reset_index(drop=True), vocab


# ----------------------------------------------------------------------------
# 2. Graph construction + SNA measures
# ----------------------------------------------------------------------------

# Aggregating evidence items into weighted bipartite edges. weight = sum of evidential status; n_items = independent item count.
def edge_table(items: pd.DataFrame) -> pd.DataFrame:
    g = (items.groupby(["Capability", "Dimension"])
              .agg(weight=("Status", "sum"), n_items=("Status", "size"))
              .reset_index())
    g["log_weight"] = g["weight"].apply(lambda w: math.log(1 + w))
    return g


# Degree (distinct partners) and strength (summed weights) per node, for both node sets of the bipartite graph.
def node_metrics(edges: pd.DataFrame, weight_col: str = "weight") -> pd.DataFrame:
    recs = []
    for side, other in (("Capability", "Dimension"), ("Dimension", "Capability")):
        grp = edges.groupby(side).agg(
            degree=(other, "nunique"),
            strength=(weight_col, "sum"),
            n_items=("n_items", "sum"),
        ).reset_index().rename(columns={side: "node"})
        grp["side"] = "capability" if side == "Capability" else "dimension"
        recs.append(grp)
    return pd.concat(recs, ignore_index=True)


# Opsahl et al. (2010) Generalised degree for capability nodes: C(alpha) = k^(1-alpha) * s^alpha  (k = degree, s = strength).
def opsahl(edges: pd.DataFrame, alpha: float) -> pd.Series:
    m = node_metrics(edges)
    m = m[m["side"] == "capability"].set_index("node")
    return (m["degree"] ** (1 - alpha)) * (m["strength"] ** alpha)


# Spearman rho between two {node: score} dicts over their common keys, with average ranks for ties
def spearman(a: dict, b: dict) -> float:
    keys = sorted(set(a) & set(b))
    if len(keys) < 3:
        return float("nan")

    def ranks(vals):
        order = sorted(range(len(vals)), key=lambda i: vals[i])
        r = [0.0] * len(vals)
        i = 0
        while i < len(vals):
            j = i
            while j + 1 < len(vals) and vals[order[j + 1]] == vals[order[i]]:
                j += 1
            avg = (i + j) / 2 + 1
            for k in range(i, j + 1):
                r[order[k]] = avg
            i = j + 1
        return r

    ra = ranks([a[k] for k in keys])
    rb = ranks([b[k] for k in keys])
    n = len(keys)
    ma, mb = sum(ra) / n, sum(rb) / n
    cov = sum((x - ma) * (y - mb) for x, y in zip(ra, rb))
    va = math.sqrt(sum((x - ma) ** 2 for x in ra))
    vb = math.sqrt(sum((y - mb) ** 2 for y in rb))
    return cov / (va * vb) if va and vb else float("nan")

# weighted betweenness and greedy-modularity communities on the weighted bipartite graph (approximation of Barber 2007) - Requires networkx
def try_networkx_extras(edges: pd.DataFrame):
    try:
        import networkx as nx
    except ImportError:
        print("NOTE: networkx not installed -> skipping betweenness and "
              "community detection (pip install networkx to enable).")
        return None, None
    G = nx.Graph()
    for _, r in edges.iterrows():
        G.add_edge(r["Capability"], r["Dimension"], weight=float(r["weight"]))
    # betweenness on inverse weights (strong ties = short distances)
    for u, v, d in G.edges(data=True):
        d["dist"] = 1.0 / d["weight"]
    btw = nx.betweenness_centrality(G, weight="dist", normalized=True)
    try:
        comms = nx.algorithms.community.greedy_modularity_communities(
            G, weight="weight")
        comms = [sorted(c) for c in comms]
    except Exception as exc:                       # pragma: no cover
        print(f"NOTE: community detection failed ({exc}).")
        comms = None
    return btw, comms


# ----------------------------------------------------------------------------
# 3. The five sensitivity passes (s3.9)
# ----------------------------------------------------------------------------

def run_passes(items: pd.DataFrame):
    results = {}          # pass label -> {capability: score}
    tables = {}           # pass label -> node metrics DataFrame

    def cap_strength(edge_df):
        m = node_metrics(edge_df)
        return dict(m[m["side"] == "capability"]
                    .set_index("node")["strength"])

    # Pass A — full weighted graph
    eA = edge_table(items)
    results["A_full_weighted"] = cap_strength(eA)
    tables["A_full_weighted"] = node_metrics(eA)

    # Pass B — empirical-only (drop evidential status 1)
    eB = edge_table(items[items["Status"] >= 2])
    results["B_empirical_only"] = cap_strength(eB)
    tables["B_empirical_only"] = node_metrics(eB)

    # Pass C — binary graph (every edge weight = 1)
    eC = eA.copy()
    eC["weight"] = 1
    results["C_binary"] = cap_strength(eC)
    tables["C_binary"] = node_metrics(eC)

    # Pass D — one graph per source type
    for stype, sub in items.groupby(items["SourceType"].astype(str).str.strip()):
        if not stype or stype.lower() == "nan":
            continue
        eD = edge_table(sub)
        label = "D_" + re.sub(r"[^A-Za-z0-9]+", "_", stype).strip("_")
        results[label] = cap_strength(eD)
        tables[label] = node_metrics(eD)

    # Pass E — Opsahl alpha sweep on the full graph
    for a in OPSAHL_ALPHAS:
        results[f"E_opsahl_a{a:g}"] = dict(opsahl(eA, a))

    return eA, results, tables


# ----------------------------------------------------------------------------
# 4. Interactive visualisation (self-contained HTML)
# ----------------------------------------------------------------------------


# Map publication year -> HSL lightness %. Older = darker, newer = brighter.
def year_lightness(year, ymin, ymax):
    if year is None or (isinstance(year, float) and math.isnan(year)):
        year = ymin
    if ymax == ymin:
        return 46
    t = (year - ymin) / (ymax - ymin)
    return round(24 + t * (68 - 24), 1)


def build_viz_payload(items, edges, vocab, results, btw, comms):
    yrs = items["Year"].dropna()
    ymin = int(yrs.min()) if len(yrs) else 2000
    ymax = int(yrs.max()) if len(yrs) else datetime.now().year

    strengthA = results["A_full_weighted"]
    m = node_metrics(edges)
    dim_strength = dict(m[m["side"] == "dimension"].set_index("node")["strength"])

    comm_of = {}
    if comms:
        for i, c in enumerate(comms):
            for n in c:
                comm_of[n] = i

    nodes = []
    for cap, meta in sorted(vocab["cap"].items()):
        if cap not in set(items["Capability"]):
            continue
        nodes.append({
            "id": cap, "side": "capability", "name": meta["name"],
            "definition": meta["definition"],
            "strength": float(strengthA.get(cap, 0)),
            "betweenness": round(float(btw.get(cap, 0)), 4) if btw else None,
            "community": comm_of.get(cap),
        })
    for dim, meta in vocab["dim"].items():
        nodes.append({
            "id": dim, "side": "dimension", "name": meta["name"],
            "definition": meta["definition"],
            "strength": float(dim_strength.get(dim, 0)),
            "betweenness": round(float(btw.get(dim, 0)), 4) if btw else None,
            "community": comm_of.get(dim),
        })

    ev_list = []
    for _, r in items.iterrows():
        year = None if pd.isna(r["Year"]) else int(r["Year"])
        ev_list.append({
            "id": r["Evidence_ID"], "src": r["Source_ID"],
            "cap": r["Capability"], "dim": r["Dimension"],
            "fine": str(r["EdgeFine"]), "group": str(r["EdgeGroup"]),
            "status": int(r["Status"]), "year": year,
            "stype": str(r["SourceType"] or ""),
            "uk": str(r["UK"] or ""), "conf": str(r["Confidence"] or ""),
            "loc": str(r["Locator"] or ""),
            "quote": str(r["Quote"] or ""),
            "mech": str(r["Mechanism"] or ""),
            "target": str(r["Target"] or ""),
            "corrob": str(r["Corrob"] or ""),
            "cite": str(r["Citation"] or ""),
            "hue": GROUP_HUES.get(str(r["EdgeGroup"]).strip(), 210),
            "light": year_lightness(year, ymin, ymax),
        })

    # top-10 ranking + Spearman matrix for the results panel in the HTML
    rankA = sorted(strengthA.items(), key=lambda kv: -kv[1])
    labels = list(results.keys())
    def _rho(la, lb):
        r = spearman(results[la], results[lb])
        return None if math.isnan(r) else round(r, 3)

    rho = [[None if i > j else _rho(la, lb) for j, lb in enumerate(labels)]
           for i, la in enumerate(labels)]

    return {
        "meta": {"ymin": ymin, "ymax": ymax,
                 "generated": datetime.now().strftime("%Y-%m-%d %H:%M"),
                 "n_items": len(ev_list),
                 "groups": list(GROUP_HUES.keys()),
                 "group_hues": GROUP_HUES,
                 "stypes": sorted({e["stype"] for e in ev_list if e["stype"]})},
        "nodes": nodes,
        "evidence": ev_list,
        "capNames": {c: v["name"] for c, v in vocab["cap"].items()},
        "dimNames": {d: v["name"] for d, v in vocab["dim"].items()},
        "rankingA": [[c, round(s, 1)] for c, s in rankA],
        "spearman": {"labels": labels, "rho": rho},
        "communities": comms if comms else [],
    }


def write_html(payload: dict, out_path: str):
    page = HTML_TEMPLATE.replace(
        "__DATA__",
        json.dumps(payload,
                   default=lambda o: o.item() if hasattr(o, "item") else str(o)))
    with open(out_path, "w", encoding="utf-8") as fh:
        fh.write(page)


# The template is kept at the bottom of the file so the analysis logic above stays readable. SVG + vanilla JavaScript: no internet access or external library is required.


# ----------------------------------------------------------------------------
# 5. Main
# ----------------------------------------------------------------------------

def main():
    wb_path = find_workbook()
    print(f"Workbook: {os.path.basename(wb_path)}")
    items, vocab = load_data(wb_path)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    for stale in glob.glob(os.path.join(OUTPUT_DIR, 'nodes_*.csv')):
        os.remove(stale)          # clear per-pass files from earlier runs

    edges, results, tables = run_passes(items)

    # log-scaling trigger (protocol s2.3)
    heavy = edges[edges["n_items"] > LOG_SCALE_TRIGGER]
    if len(heavy):
        print(f"\nNOTE: {len(heavy)} edge(s) exceed {LOG_SCALE_TRIGGER} "
              f"independent items; per protocol s2.3 the log-scaled weight "
              f"w = log(1 + sum(e_i)) is also reported in edges.csv:")
        for _, r in heavy.iterrows():
            print(f"   {r['Capability']} -- {r['Dimension']}: "
                  f"{r['n_items']} items, w={r['weight']}, "
                  f"log-w={r['log_weight']:.2f}")

    btw, comms = try_networkx_extras(edges)

    # console report
    print("\n=== Pass A — capability ranking by weighted degree "
          "(primary measure) ===")
    for i, (cap, s) in enumerate(
            sorted(results["A_full_weighted"].items(), key=lambda kv: -kv[1]), 1):
        name = vocab["cap"].get(cap, {}).get("name", "")
        b = f"  betweenness={btw[cap]:.3f}" if btw else ""
        print(f"{i:2d}. {cap}  w={s:>5.0f}   {name}{b}")
    if btw:
        print("    (betweenness is secondary/descriptive only: it is not "
              "stable under weight perturbation — Segarra & Ribeiro 2015)")

    labels = list(results.keys())
    print("\n=== Spearman rank correlations between passes "
          f"(rho > {ROBUSTNESS_RHO} = robust) ===")
    core = [l for l in labels if l[0] in "ABC" or l.startswith("E_")]
    for i, la in enumerate(core):
        for lb in core[i + 1:]:
            r = spearman(results[la], results[lb])
            flag = "ROBUST" if (not math.isnan(r) and r > ROBUSTNESS_RHO) else "  --  "
            print(f"  {la:22s} vs {lb:22s} rho={r:5.3f}  {flag}")
    dpasses = [l for l in labels if l.startswith("D_")]
    if dpasses:
        print("\n  Pass D (single source type) vs Pass A baseline:")
        for l in dpasses:
            r = spearman(results["A_full_weighted"], results[l])
            n = len(set(results[l]))
            print(f"  A vs {l:34s} rho={r:5.3f}  ({n} capabilities present)")

    if comms:
        print("\n=== Communities (greedy modularity on weighted bipartite "
              "graph; approximates Barber 2007) ===")
        for i, c in enumerate(comms, 1):
            named = [f"{n} ({vocab['cap'].get(n, vocab['dim'].get(n, {})).get('name','')})"
                     for n in c]
            print(f"  Community {i}: " + "; ".join(named))

    # CSV outputs
    items.to_csv(os.path.join(OUTPUT_DIR, "included_evidence_items.csv"),
                 index=False)
    edges.to_csv(os.path.join(OUTPUT_DIR, "edges.csv"), index=False)
    for label, t in tables.items():
        t.to_csv(os.path.join(OUTPUT_DIR, f"nodes_{label}.csv"), index=False)
    pd.DataFrame(
        {la: {lb: spearman(results[la], results[lb]) for lb in labels}
         for la in labels}
    ).to_csv(os.path.join(OUTPUT_DIR, "spearman_matrix.csv"))
    pd.DataFrame(results).to_csv(
        os.path.join(OUTPUT_DIR, "capability_scores_all_passes.csv"))
    if comms:
        with open(os.path.join(OUTPUT_DIR, "communities.txt"), "w") as fh:
            for i, c in enumerate(comms, 1):
                fh.write(f"Community {i}: {', '.join(c)}\n")

    # interactive visualisation
    payload = build_viz_payload(items, edges, vocab, results, btw, comms)
    html_path = os.path.join(OUTPUT_DIR, "SCCAH_network.html")
    write_html(payload, html_path)

    print(f"\nAll outputs written to {OUTPUT_DIR}")
    print(f"Open the interactive network:  {html_path}")


# ----------------------------------------------------------------------------
# HTML + JS template (self-contained; __DATA__ is replaced with the JSON payload)
# ----------------------------------------------------------------------------

HTML_TEMPLATE = r"""
<!DOCTYPE html>
<html lang='en'>
<head>
<meta charset='utf-8'>
<title>Capability x Autonomy-Harm Network</title>
<style>
  :root{
    --bg:#12141a; --panel:#1b1e27; --panel2:#232734; --ink:#e8eaf0;
    --muted:#9aa1b5; --line:#31374a; --accent:#6ea8fe;
  }
  *{box-sizing:border-box}
  body{margin:0;font:13px/1.45 'Segoe UI',system-ui,sans-serif;background:var(--bg);color:var(--ink);display:flex;height:100vh;overflow:hidden}
  #sidebar{width:305px;min-width:305px;background:var(--panel);border-right:1px solid var(--line);overflow-y:auto;padding:14px 14px 40px}
  #sidebar h1{font-size:15px;margin:0 0 2px}
  #sidebar .sub{color:var(--muted);font-size:11px;margin-bottom:12px}
  fieldset{border:1px solid var(--line);border-radius:8px;margin:0 0 12px;padding:8px 10px}
  legend{font-size:11px;letter-spacing:.06em;text-transform:uppercase;color:var(--muted);padding:0 5px}
  label.chk{display:flex;align-items:center;gap:7px;padding:3px 2px;cursor:pointer;border-radius:5px}
  label.chk:hover{background:var(--panel2)}
  label.chk input{accent-color:var(--accent)}
  .swatch{display:inline-block;width:22px;height:0;border-top:3px solid;border-radius:2px}
  .thick1{border-top-width:1px!important}.thick2{border-top-width:3px!important}.thick3{border-top-width:5px!important}
  .mini{font-size:11px;color:var(--muted)}
  .btnrow{display:flex;gap:6px;margin:2px 0 10px;flex-wrap:wrap}
  button{background:var(--panel2);border:1px solid var(--line);color:var(--ink);border-radius:6px;padding:3px 9px;font-size:11px;cursor:pointer}
  button:hover{border-color:var(--accent)}
  button.on{background:var(--accent);color:#0b0d12;border-color:var(--accent)}
  input[type=range]{width:100%;accent-color:var(--accent)}
  #yearlbl{font-variant-numeric:tabular-nums}
  #stage{flex:1;position:relative;overflow:hidden}
  svg{width:100%;height:100%;display:block;cursor:grab}
  svg.panning{cursor:grabbing}
  .nodeC circle{stroke:#0b0d12;stroke-width:1.5;cursor:pointer;transition:opacity .18s}
  .nodeC text{fill:var(--ink);font-size:12px;cursor:pointer;paint-order:stroke;stroke:#12141a;stroke-width:3px;stroke-linejoin:round}
  .nodeC .w{font-size:10px;fill:var(--muted)}
  path.ev,path.pj{fill:none;cursor:pointer;transition:opacity .15s}
  path.ev.dim,path.pj.dim{opacity:.05;pointer-events:none}
  .nodeC.dim{opacity:.15}
  #details{position:absolute;top:12px;right:12px;width:330px;max-height:calc(100% - 24px);overflow-y:auto;background:var(--panel);border:1px solid var(--line);border-radius:10px;padding:14px;display:none;box-shadow:0 8px 30px rgba(0,0,0,.5)}
  #details h2{font-size:14px;margin:0 26px 6px 0}
  #details .tag{display:inline-block;background:var(--panel2);border:1px solid var(--line);border-radius:20px;padding:1px 9px;font-size:11px;margin:0 4px 4px 0}
  #details dl{margin:8px 0 0}
  #details dt{color:var(--muted);font-size:10.5px;text-transform:uppercase;letter-spacing:.05em;margin-top:8px}
  #details dd{margin:1px 0 0}
  #closeD{position:absolute;top:8px;right:10px;background:none;border:none;color:var(--muted);font-size:16px;cursor:pointer}
  #count{position:absolute;left:14px;bottom:10px;color:var(--muted);font-size:11px;background:rgba(18,20,26,.75);padding:3px 9px;border-radius:6px}
  #zoomhint{position:absolute;right:14px;bottom:10px;color:var(--muted);font-size:11px;background:rgba(18,20,26,.75);padding:3px 9px;border-radius:6px}
  table.rho{border-collapse:collapse;font-size:10px;margin-top:6px}
  table.rho td,table.rho th{border:1px solid var(--line);padding:2px 4px;text-align:center}
  table.rho th{color:var(--muted);font-weight:normal}
  td.hi{background:#1d3a24}
  details.section{margin-bottom:12px;border:1px solid var(--line);border-radius:8px;padding:6px 10px}
  details.section summary{cursor:pointer;font-size:11px;letter-spacing:.06em;text-transform:uppercase;color:var(--muted)}
  ol.rank{padding-left:20px;margin:6px 0}
  ol.rank li{margin:2px 0}
  .detonly{display:none}
  body.detailed .detonly{display:block}
  a{color:var(--accent)}
</style>
</head>
<body>
<div id='sidebar'>
  <h1>Surveillance Capitalism capability &rarr; autonomy-harm network</h1>
  <div class='sub' id='meta'></div>

  <div class='btnrow' id='viewBtns'>
    <button id='vBip' class='on'>Bipartite network</button>
    <button id='vProj'>Capability projection</button>
    <button id='vAut'>Autonomy projection</button>
  </div>
  <label class='chk' style='margin-bottom:10px'><input type='checkbox' id='detToggle'><span>Detailed mode <span class='mini'>(IDs, rankings, sensitivity analysis)</span></span></label>

  <fieldset><legend>Edge type (superordinate) &mdash; colour</legend><div id='fGroups'></div></fieldset>

  <fieldset><legend>Evidential status &mdash; thickness</legend><div id='fStatus'></div></fieldset>

  <fieldset><legend>Publication year &mdash; brightness</legend>
    <div class='mini' id='yearlbl'></div>
    <input type='range' id='yMin'><input type='range' id='yMax'>
    <div class='mini'>brighter line = more recent study</div>
  </fieldset>

  <details class='section' id='advanced'><summary>Advanced filters</summary>
    <fieldset style='margin-top:8px'><legend>Capabilities</legend>
      <div class='btnrow'><button data-set='cap' data-mode='all'>all</button><button data-set='cap' data-mode='none'>none</button></div>
      <div id='fCaps'></div>
    </fieldset>
    <fieldset><legend>Autonomy dimensions</legend>
      <div class='btnrow'><button data-set='dim' data-mode='all'>all</button><button data-set='dim' data-mode='none'>none</button></div>
      <div id='fDims'></div>
    </fieldset>
    <fieldset><legend>Source type</legend>
      <div class='btnrow'><button data-set='sty' data-mode='all'>all</button><button data-set='sty' data-mode='none'>none</button></div>
      <div id='fStypes'></div>
    </fieldset>
  </details>

  <div class='btnrow'><button id='reset'>Reset all filters</button><button id='unfocus'>Clear focus</button><button id='zreset'>Reset zoom</button></div>

  <div class='detonly'>
    <details class='section'><summary>Pass A ranking (weighted degree)</summary><ol class='rank' id='rankA'></ol></details>
    <details class='section'><summary>Sensitivity: Spearman &rho; between passes</summary><div id='rhoTbl' style='overflow-x:auto'></div><div class='mini'>&rho; &gt; 0.85 (highlighted) = ranking robust to that weighting assumption.</div></details>
    <details class='section'><summary>Communities</summary><div id='comms' class='mini'></div></details>
  </div>
  <div class='mini'>Scroll to zoom, drag to pan &mdash; zoom in to separate overlapping lines and click individual edges. Click a node to focus it; click a line for the evidence behind it; click empty space to clear.</div>
</div>

<div id='stage'>
  <svg id='svg'><g id='viewport'></g></svg>
  <div id='count'></div>
  <div id='zoomhint'>scroll = zoom &middot; drag = pan</div>
  <div id='details'><button id='closeD'>&times;</button><div id='detBody'></div></div>
</div>

<script>
const DATA = __DATA__;
const SVGNS = 'http://www.w3.org/2000/svg';
const svg = document.getElementById('svg');
const viewport = document.getElementById('viewport');
const caps = DATA.nodes.filter(n=>n.side==='capability');
const dims = DATA.nodes.filter(n=>n.side==='dimension');
const EV = DATA.evidence;
const STATUS_W = {1:1.3, 2:2.7, 3:4.2};

/* ---------- state ---------- */
const state = {
  groups:new Set(DATA.meta.groups),
  status:new Set([1,2,3]),
  caps:new Set(caps.map(n=>n.id)),
  dims:new Set(dims.map(n=>n.id)),
  sty:new Set(DATA.meta.stypes.concat([''])),
  y0:DATA.meta.ymin, y1:DATA.meta.ymax,
  focus:null,
  mode:'bipartite',          // 'bipartite' | 'projection'
  detailed:false
};

function nameOf(id){ return DATA.capNames[id]||DATA.dimNames[id]||id; }
function lbl(id){ return state.detailed ? `${nameOf(id)} (${id})` : nameOf(id); }

function visible(e){
  const y = e.year===null ? DATA.meta.ymin : e.year;
  return state.groups.has(e.group) && state.status.has(e.status) &&
         state.caps.has(e.cap) && state.dims.has(e.dim) &&
         state.sty.has(e.stype) && y>=state.y0 && y<=state.y1;
}

/* ---------- zoom & pan (viewport transform) ---------- */
const vt = {k:1,x:0,y:0};
function applyVT(){ viewport.setAttribute('transform',`translate(${vt.x},${vt.y}) scale(${vt.k})`); }
svg.addEventListener('wheel',ev=>{
  ev.preventDefault();
  const r=svg.getBoundingClientRect();
  const mx=ev.clientX-r.left, my=ev.clientY-r.top;
  const f=Math.exp(-ev.deltaY*0.0016);
  const k2=Math.min(12,Math.max(0.4,vt.k*f)), s=k2/vt.k;
  vt.x = mx-(mx-vt.x)*s; vt.y = my-(my-vt.y)*s; vt.k=k2;
  applyVT();
},{passive:false});
let pan=null, dragged=false;
svg.addEventListener('pointerdown',ev=>{ pan={x:ev.clientX,y:ev.clientY,vx:vt.x,vy:vt.y}; dragged=false; });
window.addEventListener('pointermove',ev=>{
  if(!pan)return;
  const dx=ev.clientX-pan.x, dy=ev.clientY-pan.y;
  if(Math.hypot(dx,dy)>4){ dragged=true; svg.classList.add('panning'); }
  if(dragged){ vt.x=pan.vx+dx; vt.y=pan.vy+dy; applyVT(); }
});
window.addEventListener('pointerup',()=>{ pan=null; svg.classList.remove('panning'); setTimeout(()=>dragged=false,0); });
document.getElementById('zreset').onclick=()=>{ vt.k=1;vt.x=0;vt.y=0;applyVT(); };

/* ---------- layouts ---------- */
let W=0,H=0, posBi={}, posPrC={}, posPrD={};
function layout(){
  W = svg.clientWidth; H = svg.clientHeight;
  const padT=46, padB=30, xC=205, xD=W-215;
  caps.forEach((n,i)=>{ posBi[n.id]={x:xC, y:padT+(H-padT-padB)*(caps.length===1?.5:i/(caps.length-1))}; });
  dims.forEach((n,i)=>{ posBi[n.id]={x:xD, y:padT+(H-padT-padB)*(dims.length===1?.5:i/(dims.length-1))}; });
  const cx=W/2, cy=H/2, R=Math.min(W,H)/2-110;
  caps.forEach((n,i)=>{ const a=-Math.PI/2 + 2*Math.PI*i/caps.length;
    posPrC[n.id]={x:cx+R*Math.cos(a), y:cy+R*Math.sin(a)}; });
  dims.forEach((n,i)=>{ const a=-Math.PI/2 + 2*Math.PI*i/dims.length;
    posPrD[n.id]={x:cx+R*Math.cos(a), y:cy+R*Math.sin(a)}; });
}
function pos(id){
  if(state.mode==='projCap') return posPrC[id]||posBi[id];
  if(state.mode==='projDim') return posPrD[id]||posBi[id];
  return posBi[id];
}

/* parallel-edge fanning (bipartite): each evidence item gets its own offset so
   zooming in separates overlapping lines and every one is clickable */
const pairIdx = {};
EV.forEach(e=>{ const k=e.cap+'|'+e.dim; (pairIdx[k]=pairIdx[k]||[]).push(e.id); });
function edgePath(e){
  const a=posBi[e.cap], b=posBi[e.dim];
  const k=e.cap+'|'+e.dim, arr=pairIdx[k];
  const i=arr.indexOf(e.id), n=arr.length;
  const spread=(i-(n-1)/2) * Math.min(26, 190/Math.max(n,1));
  const mx=(a.x+b.x)/2, my=(a.y+b.y)/2;
  const dx=b.x-a.x, dy=b.y-a.y, len=Math.hypot(dx,dy)||1;
  const ox=-dy/len*spread, oy=dx/len*spread;
  return `M${a.x+14},${a.y} Q${mx+ox},${my+oy} ${b.x-14},${b.y}`;
}

/* ---------- build ---------- */
const gEdges=document.createElementNS(SVGNS,'g');
const gProj=document.createElementNS(SVGNS,'g');
const gNodes=document.createElementNS(SVGNS,'g');
viewport.appendChild(gEdges); viewport.appendChild(gProj); viewport.appendChild(gNodes);
const edgeEls={}, nodeEls={};
function colour(e){ return `hsl(${e.hue},64%,${e.light}%)`; }

function build(){
  layout();
  gEdges.innerHTML=''; gNodes.innerHTML='';
  EV.forEach(e=>{
    const p=document.createElementNS(SVGNS,'path');
    p.setAttribute('class','ev');
    p.setAttribute('d',edgePath(e));
    p.setAttribute('stroke',colour(e));
    p.setAttribute('stroke-width',STATUS_W[e.status]);
    p.setAttribute('stroke-linecap','round');
    p.addEventListener('click',ev=>{ if(dragged)return; ev.stopPropagation(); showEdge(e); });
    p.addEventListener('mouseenter',()=>{ if(!p.classList.contains('dim')) p.setAttribute('stroke-width',STATUS_W[e.status]+2); });
    p.addEventListener('mouseleave',()=>p.setAttribute('stroke-width',STATUS_W[e.status]));
    gEdges.appendChild(p); edgeEls[e.id]=p;
  });
  DATA.nodes.forEach(n=>{
    const g=document.createElementNS(SVGNS,'g');
    g.setAttribute('class','nodeC');
    const c=document.createElementNS(SVGNS,'circle');
    c.setAttribute('fill', n.side==='capability' ? '#4f7dd9' : '#d9784f');
    const t=document.createElementNS(SVGNS,'text');
    const wl=document.createElementNS(SVGNS,'text'); wl.setAttribute('class','w');
    g.appendChild(c); g.appendChild(t); g.appendChild(wl);
    g.addEventListener('click',ev=>{ if(dragged)return; ev.stopPropagation(); toggleFocus(n.id); });
    gNodes.appendChild(g); nodeEls[n.id]={g,c,t,wl,n};
  });
  refresh();
}

/* ---------- one-mode projections ----------
   Capability projection: two capabilities are linked if they bear on the
   same autonomy dimension under the CURRENT filters. Autonomy projection:
   two dimensions are linked if they are affected by the same capability.
   Link weight = sum over shared partners of min(w_A,p , w_B,p) — the
   overlap in evidence-weighted attention. */
function projection(visItems){
  const capSide = state.mode==='projCap';
  const wcd={};                      // projected node -> partner -> {w,n}
  visItems.forEach(e=>{
    const a = capSide? e.cap : e.dim, p = capSide? e.dim : e.cap;
    wcd[a]=wcd[a]||{};
    const d=wcd[a][p]=wcd[a][p]||{w:0,n:0};
    d.w+=e.status; d.n++;
  });
  const ids=Object.keys(wcd), links=[];
  for(let i=0;i<ids.length;i++)for(let j=i+1;j<ids.length;j++){
    const A=ids[i],B=ids[j]; let w=0; const shared=[];
    for(const d in wcd[A]) if(wcd[B][d]){
      const m=Math.min(wcd[A][d].w,wcd[B][d].w); w+=m;
      shared.push({p:d,a:wcd[A][d],b:wcd[B][d],m});
    }
    if(w>0) links.push({a:A,b:B,w,shared});
  }
  return links;
}
function drawProjection(visItems){
  gProj.innerHTML='';
  const links=projection(visItems);
  const wmax=Math.max(1,...links.map(l=>l.w));
  const P = state.mode==='projCap' ? posPrC : posPrD;
  const col = state.mode==='projCap' ? '#6ea8fe' : '#e0925f';
  links.forEach(l=>{
    const p=document.createElementNS(SVGNS,'path');
    const a=P[l.a], b=P[l.b];
    p.setAttribute('class','pj');
    p.setAttribute('d',`M${a.x},${a.y} L${b.x},${b.y}`);
    p.setAttribute('stroke',col);
    p.setAttribute('stroke-opacity',0.28+0.6*l.w/wmax);
    p.setAttribute('stroke-width',1+7*l.w/wmax);
    p.setAttribute('stroke-linecap','round');
    const inFocus = !state.focus || l.a===state.focus || l.b===state.focus;
    if(!inFocus) p.classList.add('dim');
    p.addEventListener('click',ev=>{ if(dragged)return; ev.stopPropagation(); showProjEdge(l); });
    p.addEventListener('mouseenter',()=>{ if(!p.classList.contains('dim')) p.setAttribute('stroke-width',3+7*l.w/wmax); });
    p.addEventListener('mouseleave',()=>p.setAttribute('stroke-width',1+7*l.w/wmax));
    gProj.appendChild(p);
  });
  return links;
}

/* ---------- refresh ---------- */
function refresh(){
  if(!Object.keys(edgeEls).length) return;   // not built yet
  const proj = state.mode!=='bipartite';
  gEdges.style.display = proj?'none':'';
  gProj.style.display  = proj?'':'none';

  const visItems=EV.filter(visible);
  const visW={}, visN={};
  let shown=0;
  EV.forEach(e=>{
    const v=visible(e);
    const inFocus = !state.focus || e.cap===state.focus || e.dim===state.focus;
    edgeEls[e.id].classList.toggle('dim', !(v&&inFocus));
    if(v){ shown++;
      visW[e.cap]=(visW[e.cap]||0)+e.status; visW[e.dim]=(visW[e.dim]||0)+e.status;
      visN[e.cap]=(visN[e.cap]||0)+1;        visN[e.dim]=(visN[e.dim]||0)+1; }
  });
  let links=[];
  if(proj) links=drawProjection(visItems);

  DATA.nodes.forEach(n=>{
    const hideSide = state.mode==='projCap' ? 'dimension'
                   : state.mode==='projDim' ? 'capability' : null;
    const el=nodeEls[n.id], hideNode = n.side===hideSide;
    el.g.style.display = hideNode?'none':'';
    if(hideNode) return;
    const p=pos(n.id);
    const w=visW[n.id]||0;
    const r=w? 6+Math.sqrt(w)*1.7 : 4;
    el.c.setAttribute('cx',p.x); el.c.setAttribute('cy',p.y); el.c.setAttribute('r',r);
    el.c.setAttribute('fill-opacity', w? .95 : .25);
    const left = proj ? (p.x < W/2) : n.side==='capability';
    el.t.setAttribute('x', left? p.x-(r+8) : p.x+(r+8));
    el.t.setAttribute('y', p.y+1);
    el.t.setAttribute('text-anchor', left?'end':'start');
    const nm = state.detailed ? `${n.name} (${n.id})` : n.name;
    el.t.textContent = nm.length>44 ? nm.slice(0,42)+'…' : nm;
    el.wl.setAttribute('x', left? p.x-(r+8) : p.x+(r+8));
    el.wl.setAttribute('y', p.y+14);
    el.wl.setAttribute('text-anchor', left?'end':'start');
    el.wl.textContent = w? `w=${w} · ${visN[n.id]} items` : '';
    let neighbour=true;
    if(state.focus && n.id!==state.focus){
      neighbour = proj
        ? links.some(l=>(l.a===state.focus&&l.b===n.id)||(l.b===state.focus&&l.a===n.id))
        : EV.some(e=>visible(e) && ((e.cap===state.focus&&e.dim===n.id)||(e.dim===state.focus&&e.cap===n.id)));
    }
    el.g.classList.toggle('dim', !neighbour && state.focus!==n.id && !!state.focus);
  });
  document.getElementById('count').textContent = proj
    ? `${links.length} ${state.mode==='projCap'?'capability':'autonomy-dimension'} links · from ${shown} visible evidence items` + (state.focus? ` · focused on ${lbl(state.focus)}` : '')
    : `${shown} / ${EV.length} evidence items shown` + (state.focus? ` · focused on ${lbl(state.focus)}` : '');
}

/* ---------- details panels ---------- */
const det=document.getElementById('details'), body=document.getElementById('detBody');
function esc(s){const d=document.createElement('div');d.textContent=s??'';return d.innerHTML;}
function open(){det.style.display='block';}
document.getElementById('closeD').onclick=()=>det.style.display='none';

function showEdge(e){
  body.innerHTML = `<h2>${state.detailed?esc(e.id)+' &mdash; ':''}evidence item</h2>
   <span class='tag' style='border-color:${colour(e)};color:${colour(e)}'>${esc(e.group)}</span>
   <span class='tag'>status ${e.status}</span>
   <span class='tag'>${e.year??'year unknown'}</span>
   <dl>
    <dt>Claim (edge)</dt><dd>${esc(lbl(e.cap))} &rarr; ${esc(lbl(e.dim))}<br><span class='mini'>${esc(e.fine)}</span></dd>
    <dt>Source</dt><dd>${state.detailed?esc(e.src)+' — ':''}${esc(e.cite)||esc(e.src)}</dd>
    <dt>Quote / paraphrase</dt><dd>${esc(e.quote)||'—'}</dd>
    <dt>Harm mechanism</dt><dd>${esc(e.mech)||'—'}</dd>
    <dt>Locator</dt><dd>${esc(e.loc)||'—'}</dd>
    <dt>Source type · UK relevance · coder confidence</dt>
      <dd>${esc(e.stype)||'—'} · ${esc(e.uk)||'—'} · ${esc(e.conf)||'—'}</dd>
    <dt>Target / affected party</dt><dd>${esc(e.target)||'—'}</dd>
    <dt>Corroborating sources (do not add weight)</dt><dd>${esc(e.corrob)||'—'}</dd>
   </dl>`;
  open();
}

function showProjEdge(l){
  const capSide = state.mode==='projCap';
  const rows=l.shared.sort((x,y)=>y.m-x.m).map(s=>
    `<li><b>${esc(lbl(s.p))}</b> — overlap ${s.m}<br><span class='mini'>${esc(nameOf(l.a))}: w=${s.a.w} (${s.a.n} items) · ${esc(nameOf(l.b))}: w=${s.b.w} (${s.b.n} items)</span></li>`).join('');
  body.innerHTML = `<h2>${capSide?'Capability':'Autonomy-dimension'} co-occurrence</h2>
    <span class='tag'>projection link</span><span class='tag'>weight ${l.w}</span>
    <dl>
     <dt>Between</dt><dd>${esc(lbl(l.a))} &harr; ${esc(lbl(l.b))}</dd>
     <dt>How this link is computed</dt><dd class='mini'>${capSide
       ? 'Two capabilities are linked when they bear on the same autonomy dimension under the current filters; the weight sums, over shared dimensions, the smaller of the two capabilities&apos; evidence weights on that dimension.'
       : 'Two autonomy dimensions are linked when they are affected by the same capability under the current filters; the weight sums, over shared capabilities, the smaller of the two dimensions&apos; evidence weights from that capability.'}</dd>
     <dt>${capSide?'Shared autonomy dimensions':'Shared capabilities'}</dt><dd><ul style='margin:2px 0;padding-left:18px'>${rows}</ul></dd>
    </dl>`;
  open();
}

function toggleFocus(id){
  state.focus = state.focus===id ? null : id;
  if(state.focus) showNode(id); else det.style.display='none';
  refresh();
}
function showNode(id){
  const n=DATA.nodes.find(n=>n.id===id);
  const items=EV.filter(e=>visible(e)&&(e.cap===id||e.dim===id));
  const byStat={1:0,2:0,3:0}, byGrp={}, partners={};
  let w=0;
  items.forEach(e=>{w+=e.status;byStat[e.status]++;byGrp[e.group]=(byGrp[e.group]||0)+1;
    const p=e.cap===id?e.dim:e.cap;(partners[p]=partners[p]||[]).push(e);});
  const pl=Object.entries(partners).sort((a,b)=>b[1].reduce((s,e)=>s+e.status,0)-a[1].reduce((s,e)=>s+e.status,0))
    .map(([p,es])=>`<li><b>${esc(lbl(p))}</b> — w=${es.reduce((s,e)=>s+e.status,0)} (${es.length} items${state.detailed?': '+es.map(e=>esc(e.src)).join(', '):''})</li>`).join('');
  body.innerHTML = `<h2>${esc(n.name)}${state.detailed?` <span class='mini'>(${esc(n.id)})</span>`:''}</h2>
   <span class='tag'>${n.side}</span>
   ${n.community!==null&&n.community!==undefined?`<span class='tag'>community ${n.community+1}</span>`:''}
   <dl>
    <dt>Definition</dt><dd>${esc(n.definition)||'—'}</dd>
    <dt>Under current filters</dt>
    <dd>weighted degree <b>${w}</b> · ${items.length} evidence items<br>
        status 3/2/1: ${byStat[3]} / ${byStat[2]} / ${byStat[1]}<br>
        ${Object.entries(byGrp).map(([g,c])=>esc(g)+': '+c).join(' · ')||''}</dd>
    ${state.detailed?`<dt>Full graph (Pass A)</dt>
    <dd>weighted degree ${n.strength}${n.betweenness!==null?` · betweenness ${n.betweenness} <span class='mini'>(secondary measure — unstable in weights)</span>`:''}</dd>`:''}
    <dt>Connections (strongest first)</dt><dd><ul style='margin:2px 0;padding-left:18px'>${pl||'<li>none visible</li>'}</ul></dd>
   </dl>`;
  open();
}
svg.addEventListener('click',()=>{ if(dragged)return; if(state.focus){state.focus=null;det.style.display='none';refresh();} });

/* ---------- sidebar controls ---------- */
function chk(parent,label,checked,swatchStyle,onchange){
  const l=document.createElement('label');l.className='chk';
  const i=document.createElement('input');i.type='checkbox';i.checked=checked;i.onchange=()=>onchange(i.checked);
  l.appendChild(i);
  if(swatchStyle){const s=document.createElement('span');s.className='swatch '+(swatchStyle.cls||'');s.style.borderTopColor=swatchStyle.color||'transparent';if(!swatchStyle.color)s.style.borderTop='none';l.appendChild(s);}
  const t=document.createElement('span');t.innerHTML=label;l.appendChild(t);
  parent.appendChild(l); return i;
}
const boxes={cap:{},dim:{},sty:{}};
const capLabels={}, dimLabels={};
DATA.meta.groups.forEach(g=>chk(document.getElementById('fGroups'),esc(g),true,
  {color:`hsl(${DATA.meta.group_hues[g]},64%,50%)`},v=>{v?state.groups.add(g):state.groups.delete(g);refresh();}));
[3,2,1].forEach(s=>chk(document.getElementById('fStatus'),
  `status ${s} — ${s===3?'empirically demonstrated':s===2?'empirically informed':'theoretically proposed'}`,true,
  {color:'#9aa1b5',cls:'thick'+s},v=>{v?state.status.add(s):state.status.delete(s);refresh();}));
caps.forEach(n=>{ const i=chk(document.getElementById('fCaps'),'',true,null,
  v=>{v?state.caps.add(n.id):state.caps.delete(n.id);refresh();});
  boxes.cap[n.id]=i; capLabels[n.id]=i.parentElement.querySelector('span:last-child'); });
dims.forEach(n=>{ const i=chk(document.getElementById('fDims'),'',true,null,
  v=>{v?state.dims.add(n.id):state.dims.delete(n.id);refresh();});
  boxes.dim[n.id]=i; dimLabels[n.id]=i.parentElement.querySelector('span:last-child'); });
DATA.meta.stypes.forEach(s=>boxes.sty[s]=chk(document.getElementById('fStypes'),esc(s),true,null,
  v=>{v?state.sty.add(s):state.sty.delete(s);refresh();}));
function relabelFilters(){
  caps.forEach(n=>capLabels[n.id].innerHTML = state.detailed?`<b>${esc(n.id)}</b> ${esc(n.name)}`:esc(n.name));
  dims.forEach(n=>dimLabels[n.id].innerHTML = state.detailed?`<b>${esc(n.id)}</b> ${esc(n.name)}`:esc(n.name));
}
relabelFilters();

document.querySelectorAll('button[data-set]').forEach(b=>b.onclick=()=>{
  const set=b.dataset.set, on=b.dataset.mode==='all';
  Object.entries(boxes[set]).forEach(([k,i])=>{i.checked=on;
    const target = set==='cap'?state.caps:set==='dim'?state.dims:state.sty;
    on?target.add(k):target.delete(k);});
  refresh();
});

/* view + detail toggles */
const vBip=document.getElementById('vBip'), vProj=document.getElementById('vProj'), vAut=document.getElementById('vAut');
function setMode(m){
  state.mode=m;
  vBip.classList.toggle('on',m==='bipartite');
  vProj.classList.toggle('on',m==='projCap');
  vAut.classList.toggle('on',m==='projDim');
  state.focus=null; det.style.display='none';
  refresh();
}
vBip.onclick=()=>setMode('bipartite');
vProj.onclick=()=>setMode('projCap');
vAut.onclick=()=>setMode('projDim');
document.getElementById('detToggle').onchange=function(){
  state.detailed=this.checked;
  document.body.classList.toggle('detailed',state.detailed);
  relabelFilters();
  det.style.display='none';
  refresh();
};

const yMin=document.getElementById('yMin'),yMax=document.getElementById('yMax'),ylbl=document.getElementById('yearlbl');
yMin.min=yMax.min=DATA.meta.ymin; yMin.max=yMax.max=DATA.meta.ymax;
yMin.value=DATA.meta.ymin; yMax.value=DATA.meta.ymax;
function yUpd(){
  let a=+yMin.value,b=+yMax.value; if(a>b){[a,b]=[b,a];}
  state.y0=a;state.y1=b; ylbl.textContent=`showing ${a} – ${b}`; refresh();
}
yMin.oninput=yUpd; yMax.oninput=yUpd; yUpd();

document.getElementById('reset').onclick=()=>{
  state.groups=new Set(DATA.meta.groups); state.status=new Set([1,2,3]);
  state.caps=new Set(caps.map(n=>n.id)); state.dims=new Set(dims.map(n=>n.id));
  state.sty=new Set(DATA.meta.stypes.concat([''])); state.focus=null;
  yMin.value=DATA.meta.ymin; yMax.value=DATA.meta.ymax;
  document.querySelectorAll('#sidebar fieldset input[type=checkbox]').forEach(i=>i.checked=true);
  det.style.display='none'; yUpd();
};
document.getElementById('unfocus').onclick=()=>{state.focus=null;det.style.display='none';refresh();};

/* ---------- results panels (detailed mode only) ---------- */
document.getElementById('meta').textContent =
  `${DATA.meta.n_items} included evidence items · years ${DATA.meta.ymin}–${DATA.meta.ymax} · generated ${DATA.meta.generated}`;
document.getElementById('rankA').innerHTML =
  DATA.rankingA.map(([c,s])=>`<li><b>${esc(c)}</b> ${esc(DATA.capNames[c]||'')} — w=${s}</li>`).join('');
(function(){
  const {labels,rho}=DATA.spearman;
  const short=l=>l.replace(/^A_full_weighted$/,'A').replace(/^B_empirical_only$/,'B').replace(/^C_binary$/,'C').replace(/^D_/,'D:').replace(/^E_opsahl_a/,'E α=');
  let h='<table class="rho"><tr><th></th>'+labels.map(l=>`<th>${esc(short(l))}</th>`).join('')+'</tr>';
  labels.forEach((la,i)=>{h+=`<tr><th>${esc(short(la))}</th>`+labels.map((lb,j)=>{
    const v=rho[i]&&rho[i][j]; return `<td class='${v!==null&&v>0.85?'hi':''}'>${v===null||v===undefined?'':v}</td>`;}).join('')+'</tr>';});
  document.getElementById('rhoTbl').innerHTML=h+'</table>';
})();
document.getElementById('comms').innerHTML = DATA.communities.length
  ? DATA.communities.map((c,i)=>`<p><b>Community ${i+1}:</b> ${c.map(id=>esc(lbl(id))).join(', ')}</p>`).join('')
  : 'Community detection requires networkx (pip install networkx), or produced no result.';

window.addEventListener('resize',()=>{layout();EV.forEach(e=>edgeEls[e.id].setAttribute('d',edgePath(e)));refresh();});
build();
</script>
</body>
</html>

"""


if __name__ == "__main__":
    main()
